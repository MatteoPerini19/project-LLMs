"""
audit_fuzzy_matches.py
======================
Generate an Excel report that highlights **near‑duplicate English items**
whose translations diverge across languages.

Public API
----------
run_audit(df, *, lang_cols=None, threshold=0.92) -> Path
    • *df* is the fully translated DataFrame produced by the orchestrator.  
    • *lang_cols* is an optional list of language column names to check
      (defaults to every column except 'en' and 'cell_id').  
    • *threshold* is the RapidFuzz similarity cut‑off (0‑1 scale).

The report is written to  ``outputs/reports/ambiguous_items.xlsx`` and a
Path object is returned to the caller for convenience.
"""

from __future__ import annotations

import itertools
from pathlib import Path
from typing import Iterable, List, Tuple

import openpyxl
import pandas as pd
from openpyxl.styles import PatternFill
from rapidfuzz.fuzz import ratio

from config import REPORT_DIR
from utils import normalize_whitespace, rapidfuzz_ratio, safe_mkdirs

# --------------------------------------------------------------------------- #
# Internal helpers
# --------------------------------------------------------------------------- #


def _similar_pairs(texts: List[str], threshold: float) -> List[Tuple[int, int]]:
    """
    Return index pairs (i, j) where texts[i] and texts[j] are near‑duplicates.
    """
    pairs = []
    for (i, a), (j, b) in itertools.combinations(enumerate(texts), 2):
        if rapidfuzz_ratio(a, b) >= threshold and a != b:
            pairs.append((i, j))
    return pairs


def _compare_translations(row_i, row_j, lang_cols) -> List[Tuple[str, str, str, str]]:
    """
    For each language column, return (lang, t_i, t_j, status).
    """
    results = []
    for lang in lang_cols:
        t_i = normalize_whitespace(str(row_i[lang]))
        t_j = normalize_whitespace(str(row_j[lang]))
        status = "OK" if t_i == t_j else "MISMATCH"
        results.append((lang, t_i, t_j, status))
    return results


# --------------------------------------------------------------------------- #
# Public entry‑point
# --------------------------------------------------------------------------- #


def run_audit(
    df: pd.DataFrame,
    *,
    lang_cols: Iterable[str] | None = None,
    threshold: float = 0.92,
) -> Path:
    """
    Scan *df* for near‑duplicate English strings with divergent translations
    and write an Excel report.

    Returns
    -------
    Path
        Location of the generated XLSX file.
    """
    if "en" not in df.columns:
        raise ValueError("DataFrame must contain an 'en' column with English text.")

    lang_cols = (
        [c for c in df.columns if c not in ("en", "cell_id")]
        if lang_cols is None
        else list(lang_cols)
    )

    en_list = df["en"].astype(str).tolist()
    near_pairs = _similar_pairs(en_list, threshold)

    records = []
    for i, j in near_pairs:
        row_i = df.iloc[i]
        row_j = df.iloc[j]
        for lang, t_i, t_j, status in _compare_translations(row_i, row_j, lang_cols):
            if status == "OK":
                continue  # skip clean pairs to keep report lean
            records.append(
                {
                    "EN #1": row_i["en"],
                    "EN #2": row_j["en"],
                    "Language": lang,
                    "T #1": t_i,
                    "T #2": t_j,
                    "Status": status,
                }
            )

    if not records:
        # No mismatches – create a minimal file stating so
        records.append(
            {
                "EN #1": "—",
                "EN #2": "—",
                "Language": "—",
                "T #1": "—",
                "T #2": "—",
                "Status": "No mismatches detected",
            }
        )

    report_df = pd.DataFrame.from_records(records)

    safe_mkdirs(REPORT_DIR)
    out_path = REPORT_DIR / "ambiguous_items.xlsx"
    with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
        report_df.to_excel(writer, index=False, sheet_name="Mismatches")

    # Apply rudimentary color coding
    wb = openpyxl.load_workbook(out_path)
    ws = wb.active
    red = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    green = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")

    status_col = [cell.value for cell in ws[1]].index("Status") + 1
    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=status_col)
        if cell.value == "MISMATCH":
            cell.fill = red
        elif cell.value == "OK":
            cell.fill = green

    wb.save(out_path)
    return out_path


__all__ = ["run_audit"]
